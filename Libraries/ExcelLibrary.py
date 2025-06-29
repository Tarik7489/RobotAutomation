import openpyxl
import pandas as pd

class ExcelLibrary:
    def read_excel_data(self, file_path, sheet_name):
        """Reads an Excel file and returns data as a list of dictionaries."""
        df = pd.read_excel(file_path, sheet_name=sheet_name, engine="openpyxl")
        return df.to_dict(orient="records")

    def write_excel_data(self, file_path, sheet_name, column_name, new_password):
        """Writes the new password into the given Excel file and sheet."""
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]

        # Find column index based on column_name
        column_index = None
        for col_num, col_cell in enumerate(sheet[1], start=1):
            if col_cell.value == column_name:
                column_index = col_num
                break

        if column_index is None:
            raise ValueError(f"Column '{column_name}' not found in {sheet_name}")

        # Write to the next empty row in that column
        for row_num in range(2, sheet.max_row + 2):
            if sheet.cell(row=row_num, column=column_index).value is None:
                sheet.cell(row=row_num, column=column_index, value=new_password)
                break

        workbook.save(file_path)
        workbook.close()

